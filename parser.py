import copy

import openpyxl
import pandas as pd

import colors
from colors import YELLOW, ORANGE, BLUE


def load_file(path):
    # Загрузка файла Excel
    workbook = openpyxl.load_workbook(path)
    sheet = workbook.active
    return sheet


def generate_table_ranges(sheet) -> list:
    # Текст, который нужно найти
    table_end = 'Часы факт. отраб'
    results_tables = []
    table_form = {
        "№ п/п": None,
        "Ф.И.О.": None,
        "Профессия (должность)": None,
        "Дни факт. отраб": None,
        "Часы факт. отраб": None,
        "Line": None
    }

    for row in sheet.iter_rows(min_row=12):
        for cell in row:
            if cell.value in table_form.keys():
                if cell.value == table_end:
                    table_form["Line"] = cell.row + 2
                    table_form["Days"] = [table_form["Профессия (должность)"] + 1, table_form["Дни факт. отраб"]]
                    table_form[cell.value] = cell.column
                    results_tables.append(copy.deepcopy(table_form))
                    table_form["Days"] = []
                else:
                    table_form[cell.value] = cell.column

    return results_tables


# Функция для получения индекса колонки по имени
def get_column_index(column_letter):
    return openpyxl.utils.column_index_from_string(column_letter)


def iterate_over_table_ranges(sheet=None, count_per_month: bool = False):
    results_tables = generate_table_ranges(sheet=sheet)
    final_matrix = [["Ф.И.О.", "Профессия (должность)", "Желтый", "Оранжевый","Синий"]]
    result_dict = {}
    if not count_per_month:
        for table in results_tables:
            line = table["Line"]
            prof = table["Профессия (должность)"]
            while sheet.cell(row=line, column=prof).value:
                for k, v in table.items():
                    if k == "Ф.И.О.":
                        fio_cell = sheet.cell(row=line, column=v).value
                    if k == "Профессия (должность)":
                        prof_cell = sheet.cell(row=line, column=v).value
                        distinct_cell = fio_cell + ';' + prof_cell
                        try:
                            result_dict[distinct_cell]
                        except KeyError:
                            result_dict[distinct_cell] = {"Yellow": 0,
                                                          "Blue": 0,
                                                          "Orange": 0}
                    elif k == "Days" and v:
                        for day in range(v[0], v[1]):
                            cell = sheet.cell(row=line, column=day).fill.fgColor.rgb
                            if cell == YELLOW:
                                result_dict[distinct_cell]['Yellow'] += 1
                            elif cell == ORANGE:
                                result_dict[distinct_cell]['Orange'] += 1
                            elif cell == BLUE:
                                result_dict[distinct_cell]['Blue'] += 1
                line += 1
        for k, v in result_dict.items():
            k = k.split(";")
            final_matrix.append([k[0], k[1], v['Yellow'], v['Orange'],v['Blue']])
        return final_matrix


def create_report(data):
    df = pd.DataFrame(data)
    styled_df = df.style.map(lambda x: colors.header_colours(x) if isinstance(x, str) else '',
                             subset=pd.IndexSlice[0, :])
    return df
