import copy

import openpyxl
import pandas as pd
from openpyxl.styles import Border, Font, Side

import colors
from colors import YELLOW, ORANGE, BLUE


def load_file(path):
    # Загрузка файла Excel
    workbook = openpyxl.load_workbook(path)
    sheet = workbook.active
    return sheet


def generate_table_ranges(sheet) -> list:
    # Текст, который нужно найти
    table_end = 'Часы факт. отраб'
    results_tables = []
    table_form = {
        "№ п/п": None,
        "Ф.И.О.": None,
        "Профессия (должность)": None,
        "Дни факт. отраб": None,
        "Часы факт. отраб": None,
        "Line": None
    }

    for row in sheet.iter_rows(min_row=12):
        for cell in row:
            if cell.value in table_form.keys():
                if cell.value == table_end:
                    table_form["Line"] = cell.row + 2
                    table_form["Days"] = [table_form["Профессия (должность)"] + 1, table_form["Дни факт. отраб"]]
                    table_form[cell.value] = cell.column
                    results_tables.append(copy.deepcopy(table_form))
                    table_form["Days"] = []
                else:
                    table_form[cell.value] = cell.column

    return results_tables


# Функция для получения индекса колонки по имени
def get_column_index(column_letter):
    return openpyxl.utils.column_index_from_string(column_letter)


def iterate_over_table_ranges(sheet=None, count_per_month: bool = False):
    results_tables = generate_table_ranges(sheet=sheet)
    final_matrix = [["Ф.И.О.", "Профессия (должность)", "Желтый", "Оранжевый", "Синий"]]
    result_dict = {}
    if not count_per_month:
        for table in results_tables:
            line = table["Line"]
            prof = table["Профессия (должность)"]
            while sheet.cell(row=line, column=prof).value:
                for k, v in table.items():
                    if k == "Ф.И.О.":
                        fio_cell = sheet.cell(row=line, column=v).value
                    if k == "Профессия (должность)":
                        prof_cell = sheet.cell(row=line, column=v).value
                        distinct_cell = fio_cell + ';' + prof_cell
                        try:
                            result_dict[distinct_cell]
                        except KeyError:
                            result_dict[distinct_cell] = {"Yellow": 0,
                                                          "Blue": 0,
                                                          "Orange": 0}
                    elif k == "Days" and v:
                        for day in range(v[0], v[1]):
                            cell = sheet.cell(row=line, column=day).fill.fgColor.rgb
                            if cell == YELLOW:
                                result_dict[distinct_cell]['Yellow'] += 1
                            elif cell == ORANGE:
                                result_dict[distinct_cell]['Orange'] += 1
                            elif cell == BLUE:
                                result_dict[distinct_cell]['Blue'] += 1
                line += 1
        for k, v in result_dict.items():
            k = k.split(";")
            final_matrix.append([k[0], k[1], v['Yellow'], v['Orange'], v['Blue']])
        return final_matrix


def create_report(data):
    df = pd.DataFrame(data)
    styled_df = df.style.map(lambda x: colors.header_colours(x) if isinstance(x, str) else '',
                             subset=pd.IndexSlice[0, :])
    return df


def colorize_and_format_file(iterated_file, file_name: str):
    # Установка максимальной ширины столбца
    pd.set_option('max_colwidth', 300)
    # Сохранение DataFrame в CSV
    with pd.ExcelWriter(f"{file_name}", engine='openpyxl') as writer:
        iterated_file.to_excel(writer, sheet_name='Report', index=False, header=False)
        worksheet = writer.sheets['Report']
        # Применяем стили к заголовкам
        bold_font = Font(bold=True)
        thin_border = Border(left=Side(style='thin'),
                             right=Side(style='thin'),
                             top=Side(style='thin'),
                             bottom=Side(style='thin'))

        # Применяем цвет к заголовкам (первая строка)
        for col in worksheet.iter_cols(min_row=1, max_row=1):  # Только первая строка
            for cell in col:
                colour = colors.header_colours(cell.value)
                cell.fill = colour
                cell.font = bold_font  # Применяем жирный шрифт к заголовкам
                cell.border = thin_border  # Применяем границы к заголовкам

        # Устанавливаем ширину для первых двух столбцов
        for column in worksheet.iter_cols(min_row=1, max_row=1):
            column_letter = column[0].column_letter
            if column[0].column <= 2:  # Для первых двух столбцов
                worksheet.column_dimensions[column_letter].width = 40
            else:  # Для остальных столбцов
                worksheet.column_dimensions[column_letter].width = 30
