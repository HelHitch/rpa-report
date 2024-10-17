from openpyxl.styles import PatternFill

YELLOW = 'FFFFFF00'
ORANGE = 'FFFFC000'
BLUE = 'FFCCCCFF'


def header_colours(header):
    if header in ('Yellow', 'Желтый'):
        return PatternFill(start_color=YELLOW, fill_type='solid')  # Желтый
    elif header in ('Orange', 'Оранжевый'):
        return PatternFill(start_color=ORANGE, fill_type='solid')  # Оранжевый
    elif header in ('Blue', 'Синий'):
        return PatternFill(start_color=BLUE, fill_type='solid')  # Синий
    else:
        return PatternFill(fgColor="FFFFFFFF", fill_type='solid')  # Цвет по умолчанию для остальных
