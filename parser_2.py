import copy

import openpyxl
import pandas as pd

import colors
from colors import YELLOW, ORANGE, BLUE


def load_file(path):
    # Загрузка файла Excel
    workbook = openpyxl.load_workbook(path)
    sheet = workbook.active
    return sheet


def generate_table_ranges(sheet) -> list:
    # Текст, который нужно найти
    table_end = 'Часы факт. отраб'
    results_tables = []
    table_form = {
        "№ п/п": None,
        "Ф.И.О.": None,
        "Профессия (должность)": None,
        "Дни факт. отраб": None,
        "Часы факт. отраб": None,
        "Line": None
    }

    for row in sheet.iter_rows(min_row=12):
        for cell in row:
            if cell.value in table_form.keys():
                if cell.value == table_end:
                    table_form["Line"] = cell.row + 2
                    table_form["Days"] = [table_form["Профессия (должность)"] + 1, table_form["Дни факт. отраб"]]
                    table_form[cell.value] = cell.column
                    results_tables.append(copy.deepcopy(table_form))
                    table_form["Days"] = []
                else:
                    table_form[cell.value] = cell.column

    return results_tables


# Функция для получения индекса колонки по имени
def get_column_index(column_letter):
    return openpyxl.utils.column_index_from_string(column_letter)


def iterate_over_table_ranges(sheet=None, count_per_month: bool = False):
    results_tables = generate_table_ranges(sheet=sheet)
    result_matrix = [["FIO", "JOB", "Yellow", "Blue", "Orange"]]
    if not count_per_month:
        for table in results_tables:
            line = table["Line"]
            prof = table["Профессия (должность)"]
            while sheet.cell(row=line, column=prof).value:
                results = [None, None, None, None, None]
                yellow = 0
                blue = 0
                orange = 0
                for k, v in table.items():
                    if k == "Ф.И.О.":
                        fio_cell = sheet.cell(row=line, column=v).value
                        results[0] = fio_cell
                    if k == "Профессия (должность)":
                        prof_cell = sheet.cell(row=line, column=v).value
                        results[1] = prof_cell
                    elif k == "Days" and v:
                        for day in range(v[0], v[1]):
                            cell = sheet.cell(row=line, column=day).fill.fgColor.rgb
                            if cell == YELLOW:
                                yellow += 1
                            elif cell == ORANGE:
                                orange += 1
                            elif cell == BLUE:
                                blue += 1
                        results[2] = yellow
                        results[4] = orange
                        results[3] = blue
                        result_matrix.append(results)
                line += 1
        return result_matrix


def create_report(data):
    # df = pd.DataFrame(data).T
    df = pd.DataFrame(data)
    styled_df = df.style.map(lambda x: colors.header_colours(x) if isinstance(x, str) else '',
                             subset=pd.IndexSlice[0, :])
    return df
